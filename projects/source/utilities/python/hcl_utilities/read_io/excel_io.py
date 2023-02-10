# Functions, methods and utilities for reading datasets from Excel documents.
# Concept: Excel datasets I/O related functions.
# Project: Historic Coastal Landfills
# Organisation: A Future Without Rubbish CIC, UK
# Author: Bharadwaj Raman
# Date First Authored: 04/02/2023

import pathlib
import typing

import openpyxl
import pandas

from hcl_constants.constants import QUALIFIED_DATASET_FILE, USEFUL_COLS, logger


class ExcelLoadWBContextManagerSupported:
    """Python `with` Context manager supported Excel load workbook into file pointer class and methods."""

    # TODO: Change this class definition to be automatically adaptive instead of having to manually update this based
    #  -- on the changes in `openpyxl.load_workbook()` method.
    def __init__(
        self,
        filename: pathlib.Path,
        read_only: bool = False,
        keep_vba: bool = False,
        data_only: bool = False,
        keep_links: bool = True,
    ):
        self.filename: pathlib.Path = filename
        self.read_only: bool = read_only
        self.keep_vba: bool = keep_vba
        self.data_only: bool = data_only
        self.keep_links: bool = keep_links

        self.workbook: typing.Optional[openpyxl.workbook.Workbook] = None

    def __enter__(self) -> openpyxl.workbook.Workbook:
        self.workbook = self.load_workbook()
        return self.workbook

    def __exit__(self, *args):
        self.workbook.close()
        self.workbook = None

    def load_workbook(self) -> openpyxl.workbook.Workbook:
        return openpyxl.load_workbook(
            self.filename,
            self.read_only,
            self.keep_vba,
            self.data_only,
            self.keep_links,
        )


def load_excel_column_headers(
    dataset_path: pathlib.Path = QUALIFIED_DATASET_FILE,
    sheet_index: int = 0,
) -> tuple[list[str], list[str], list[int]]:
    """Load Excel column headers from the dataset and their corresponding column names and indices."""
    logger.info(f"Reading initial dataset file: {dataset_path}")
    logger.info(
        f"Converting useful column header names to Excel column letters and indices."
    )
    with ExcelLoadWBContextManagerSupported(dataset_path, read_only=True) as workbook:
        worksheet = workbook.worksheets[sheet_index]
        col_headers = []
        col_letters = []
        col_indices = []

        for cell in worksheet[1]:
            col_headers.append(cell.value)
            col_letters.append(cell.column_letter)
            col_indices.append(
                cell.column - 1
            )  # We need zero-based indexing for pandas DataFrames

    return col_headers, col_letters, col_indices


def convert_useful_col_names_to_col_letters_and_indices(
    col_headers: list[str], col_letters: list[str], col_indices: list[int]
) -> tuple[list[str], list[int]]:
    """Convert the predefined useful column names to their corresponding Excel column letters and indices."""
    useful_col_letters = []
    useful_col_nums = []

    for each_col_name in USEFUL_COLS:
        useful_col_arr_index = col_headers.index(each_col_name)
        useful_col_letters.append(col_letters[useful_col_arr_index])
        useful_col_nums.append(col_indices[useful_col_arr_index])

    return useful_col_letters, useful_col_nums


def read_dataset_to_df(
    dataset_path: pathlib.Path = QUALIFIED_DATASET_FILE,
    sheet_name: str = "Sites",
    cols: list[int] = [],
) -> pandas.DataFrame:
    """Read the canonical Excel HCL site dataset into pandas DataFrame."""
    logger.info(f"Reading initial dataset file: {dataset_path}")
    # noinspection PyTypeChecker
    hld_df = pandas.read_excel(
        pathlib.Path(dataset_path), sheet_name=sheet_name, usecols=cols
    )
    assert hld_df.shape[1] == len(USEFUL_COLS)
    return hld_df


# Answer inspired by: https://stackoverflow.com/a/45312360
def convert_excel_column_index_to_column_letters(n) -> str:
    """Convert column index to Excel-style column letters, e.g., 1 = A, 26 = Z, 27 = AA, 703 = AAA."""
    name = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        name = chr(r + ord("A")) + name
    return name


# Answer inspired by: https://stackoverflow.com/a/45312360
def convert_excel_column_letters_to_column_index(name) -> int:
    """Convert column letter to Excel-style index, e.g., A = 1, Z = 26, AA = 27, AAA = 703."""
    n = 0
    for c in name:
        n = n * 26 + 1 + ord(c) - ord("A")
    return n
