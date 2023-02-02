# Read HCL canonical dataset, preprocess data, plot site markers on an interactive map.
# Concept: Plotting coordinate markers on a base Folium Map with interactive popups and layers.
# Project: Historic Coastal Landfills
# Organisation: A Future Without Rubbish CIC, UK
# Author: Bharadwaj Raman
# Date First Authored: 27/01/2023

import math
import operator
import os
import pathlib
import pickle
import time

import folium
import loguru
import numpy
import openpyxl
import pandas
from geopy.geocoders import Nominatim

import hcl_math.combinations
import hcl_math.coordinates

useful_cols = [
    "HLD reference",
    "Site name",
    "Site address",
    "Licence ./ permit reference",
    "REGIS reference",
    "WRC reference",
    "BGS reference",
    "WRA reference",
    "Permit / licence holder",
    "Permit / licence holder address",
    "Site operator's name",
    "Site operator's address",
    "Site County",
    "Country",
    "Local Authority",
    "District/Unitary Council",
    "County Council",
    "Ward",
    "Constituency",
    "Region",
    "OS prefix",
    "Easting",
    "Northing",
    "Latitude",
    "Longitude",
    "Landfill Size",
    "Landfill Notes",
    "EA Region",
    "EA Area",
    "Licence issue date",
    "Licence surrender date",
    "First input date",
    "Last input date",
    "Inert Waste",
    "Industrial Waste",
    "Commercial Waste",
    "Household Waste",
    "Special / hazardous Waste",
    "Liquid / sludge Waste",
    "Waste unknown",
    "Gas control",
    "Leachate containment",
    "Exempt",
    "Licensed",
    "No licence required",
    "Buffer point",
    "New Update CE Property Jan 2023?",
    "Borders other site(s)",
    "Site(s) nearby",
    "Is study available?",
]


CURRENT_DIR = pathlib.Path(os.getcwd())
PARENT_DATASET_PATH = CURRENT_DIR.parent.parent.parent.parent.parent / "datasets"
DATASET_FILE_NAME = "UK_Historic_Landfill_Sites.xlsx"
QUALIFIED_DATASET_FILE = PARENT_DATASET_PATH / DATASET_FILE_NAME
INTERMEDIATE_PICKLE_FILE_NAME = (
    "saved_intermediate_states/hld_df_where_on_or_adjacent_ce_property_yes.pkl"
)
QUALIFIED_INTERMEDIATE_PICKLE_FILE = CURRENT_DIR / INTERMEDIATE_PICKLE_FILE_NAME
FOLIUM_MAP_FILE_NAME = "saved_outputs/map.html"
QUALIFIED_FOLIUM_MAP_FILE = CURRENT_DIR / FOLIUM_MAP_FILE_NAME

logger = loguru.logger


def get_column_letters_from_names(
    dataset_path: pathlib.Path = QUALIFIED_DATASET_FILE,
    sheet_index: int = 0,
) -> tuple[list[str], list[str], list[int]]:
    """Get column headers from dataset and convert column names to Excel column letters and their column indices."""
    logger.info(f"Reading initial dataset file: {dataset_path}")
    logger.info(
        f"Converting useful column header names to Excel column letters and indices."
    )
    workbook = openpyxl.load_workbook(dataset_path, read_only=True)
    worksheet = workbook.worksheets[sheet_index]
    col_headers = []
    col_letters = []
    col_indices = []

    for cell in worksheet[1]:
        col_headers.append(cell.value)
        col_letters.append(cell.column_letter)
        col_indices.append(
            cell.column - 1
        )  # We need zero-based indexing for pandas DataFrames

    return col_headers, col_letters, col_indices


def convert_useful_col_names_to_col_letters_and_indices(
    col_headers: list[str], col_letters: list[str], col_indices: list[int]
) -> tuple[list[str], list[int]]:
    """Convert the predefined useful column names to their corresponding Excel column letters and indices."""
    useful_col_letters = []
    useful_col_nums = []

    for each_col_name in useful_cols:
        useful_col_arr_index = col_headers.index(each_col_name)
        useful_col_letters.append(col_letters[useful_col_arr_index])
        useful_col_nums.append(col_indices[useful_col_arr_index])

    return useful_col_letters, useful_col_nums


def read_dataset_to_df(
    dataset_path: pathlib.Path = QUALIFIED_DATASET_FILE,
    sheet_name: str = "Sites",
    cols: list[int] = [],
) -> pandas.DataFrame:
    """Read the canonical Excel HCL site dataset into pandas DataFrame."""
    logger.info(f"Reading initial dataset file: {dataset_path}")
    # noinspection PyTypeChecker
    hld_df = pandas.read_excel(
        pathlib.Path(dataset_path), sheet_name=sheet_name, usecols=cols
    )
    assert hld_df.shape[1] == len(useful_cols)
    return hld_df


def filter_dataset(
    hld_df: pandas.DataFrame,
    filter_column_name: str,
    filter_criteria: list,
    combination_operator: operator,
) -> pandas.DataFrame:
    """Apply filtration criteria on the specified columns in the DataFrame."""
    starting_shape = hld_df.shape
    logger.info(
        f"Filtering dataset with initial shape: {starting_shape}, "
        f"filter_column_name: {filter_column_name}, filter_criteria: {filter_criteria}, "
        f"combination_operator: {combination_operator.__name__}"
    )
    combination_operator(
        hld_df[filter_column_name] == filter_criteria[0],
        hld_df[filter_column_name] == filter_criteria[1],
    )

    hld_df = hld_df[
        combination_operator(
            hld_df[filter_column_name] == filter_criteria[0],
            hld_df[filter_column_name] == filter_criteria[1],
        )
    ]
    # Remove the undocumented landfill site with NAN easting and northing  # Shape(281, 50)
    hld_df = hld_df.iloc[:-1, :]

    logger.info(f"Filtered dataset shape: {hld_df.shape} out of {starting_shape}")
    return hld_df


def get_lat_long_postcode_from_easting_and_northing(
    hld_df: pandas.DataFrame,
) -> pandas.DataFrame:
    """Convert easting and northing to latitude, longitude and extract address and postcode from the coordinates."""
    latitudes = []
    longitudes = []
    postcodes = []

    for index, (each_easting, each_northing) in enumerate(
        zip(hld_df["Easting"], hld_df["Northing"])
    ):
        logger.info(
            f"Converting easting and northing into latitude and longitude for site: {index} of "
            f"{hld_df.shape[0]}"
        )
        (
            latitude,
            longitude,
        ) = hcl_math.coordinates.convert_easting_northing_to_latitude_longitude(
            each_easting, each_northing
        )
        geolocator = Nominatim(user_agent="geoapi_hcl")

        logger.info(f"Performing address lookup")
        # Get the location information
        location = geolocator.reverse(f"{latitude}, {longitude}", exactly_one=True)
        # Get the address information
        if location is not None:
            address = location.raw["address"]
            # Get the postcode
            logger.info(f"Extracting postcode from address")
            postcode = address.get("postcode", "NA")
        else:
            address = "NA"
            # Get the postcode
            postcode = "NA"

        latitudes.append(latitude)
        longitudes.append(longitude)
        postcodes.append(postcode)

    hld_df["Latitude"] = latitudes
    hld_df["Longitude"] = longitudes
    hld_df["Postcode"] = postcodes

    return hld_df


def reorder_df_columns(hld_df: pandas.DataFrame) -> pandas.DataFrame:
    """Reorder dataset pandas DataFrame for easy lookup."""
    logger.info("Reordering columns in the dataframe")
    # Columns - New Order
    new_cols_order = [
        "HLD reference",
        "Site name",
        "Site address",
        "Postcode",
        "Easting",
        "Northing",
        "Latitude",
        "Longitude",
        "Licence ./ permit reference",
        "REGIS reference",
        "WRC reference",
        "BGS reference",
        "WRA reference",
        "Permit / licence holder",
        "Permit / licence holder address",
        "Site operator's name",
        "Site operator's address",
        "Site County",
        "Country",
        "Local Authority",
        "District/Unitary Council",
        "County Council",
        "Ward",
        "Constituency",
        "Region",
        "OS prefix",
        "Landfill Size",
        "Landfill Notes",
        "EA Region",
        "EA Area",
        "Licence issue date",
        "Licence surrender date",
        "First input date",
        "Last input date",
        "Inert Waste",
        "Industrial Waste",
        "Commercial Waste",
        "Household Waste",
        "Special / hazardous Waste",
        "Liquid / sludge Waste",
        "Waste unknown",
        "Gas control",
        "Leachate containment",
        "Exempt",
        "Licensed",
        "No licence required",
        "Buffer point",
        "New Update CE Property Jan 2023?",
        "Borders other site(s)",
        "Site(s) nearby",
        "Is study available?",
    ]

    hld_df = hld_df[new_cols_order]
    return hld_df


def save_intermediate_state(
    hld_df: pandas.DataFrame,
    file_path: pathlib.Path = QUALIFIED_INTERMEDIATE_PICKLE_FILE,
):
    """Persist precious intermediate state to disk for quicker lookups."""
    logger.info(f"Saving intermediate state to file: {file_path}")
    with open(file_path, "wb") as fd:
        pickle.dump(hld_df, fd)


def read_intermediate_state(
    file_path: pathlib.Path = QUALIFIED_INTERMEDIATE_PICKLE_FILE,
) -> pandas.DataFrame:
    """Read from persisted intermediate state to speed up the analysis."""
    logger.info(f"Reading intermediate state from file: {file_path}")
    with open(file_path, "rb") as fd:
        hld_df = pickle.load(fd)
        return hld_df


def create_initial_folium_map(
    coordinates: pandas.DataFrame, tiles: str = "OpenStreetMap"
) -> folium.Map:
    """Create a base Folium map centered around the mean of the given coordinates."""
    latitude_mean = coordinates["Latitude"].mean()
    longitude_mean = coordinates["Longitude"].mean()
    logger.info(
        f"Creating the initial Folium map with tile: {tiles} centred around: {latitude_mean, longitude_mean}"
    )
    # Supported Tiles: ["OpenStreetMap", "Stamen Toner", "Stamen Terrain", "MapQuest Open Aerial", "Mapbox Bright",
    # "Mapbox Control Room", "Stamen Watercolor", "CartoDB", "CartoDB Dark Matter"]
    # Reference: https://deparkes.co.uk/2016/06/10/folium-map-tiles/
    # Reference: https://www.python-graph-gallery.com/288-map-background-with-folium
    # Reference: https://levelup.gitconnected.com/creating-interactive-maps-with-python-folium-and-some-html-f8ac716966f
    # Comparison of different tile sets: https://leaflet-extras.github.io/leaflet-providers/preview/
    folium_map = folium.Map(
        location=[latitude_mean, longitude_mean], tiles=tiles, zoom_start=5
    )

    return folium_map


def populate_each_html_table_row_popup(site_details: pandas.Series) -> str:
    """Populate a custom formatted HTML table for each site based on their details."""
    # Inspired by: https://towardsdatascience.com/folium-map-how-to-create-a-table-style-pop-up-with-html-code-76903706b88a  # noqa

    html = """
        <!DOCTYPE html>
        <html>
        <style>
        table, th, td {
          border:1px solid black;
        }
        tr:hover {background-color: cornsilk;}
        </style>
        <body>

        <h4>Landfill Site Details</h2>

        <table style="width:100%">
          <tr>
            <th style="background-color:#D6EEEE">Site Attribute Name</th>
            <th style="background-color:#D6EEEE">Site Attribute Value</th>
          </tr>
          """

    each_table_details = []
    for each_col_name, each_col_value in zip(list(site_details.index), site_details):
        if type(each_col_value) in [
            int,
            float,
            numpy.int64,
            numpy.float64,
        ] and math.isnan(each_col_value):
            each_col_value = "NA"

        tmp_str = f"""
          <tr>
            <td style="color:blue;">{each_col_name}</td>
            <td>{each_col_value}</td>
          </tr>
        """
        each_table_details.append(tmp_str)
        html += tmp_str

    html_body_rest = """
    </table>

    <p>Site details were imported from the official dataset.</p>

    </body>
    </html>
    """
    html += html_body_rest

    return html


def plot_site_markers_on_map(
    hld_df: pandas.DataFrame,
    marker_colour: str,
    marker_layer_name: str,
    folium_map: folium.Map,
) -> folium.Map:
    """Plot various site markers an already created instance of Folium Map."""

    logger.info(
        f"Plotting site markers for: {marker_layer_name} | with the colour: {marker_colour} | on the map."
    )
    # Define the marker icon style
    icon_style = "fa-solid fa-xmark"

    # Supported default Marker colours
    colours = [
        "red",
        "blue",
        "gray",
        "darkred",
        "lightred",
        "orange",
        "beige",
        "green",
        "darkgreen",
        "lightgreen",
        "darkblue",
        "lightblue",
        "purple",
        "darkpurple",
        "pink",
        "cadetblue",
        "lightgray",
        "black",
    ]

    marker_layer = folium.FeatureGroup(name=marker_layer_name)

    # Add markers for each of the given coordinates
    for site_index, site_details in hld_df.iterrows():
        styled_icon = folium.Icon(prefix="fa", icon=icon_style, color=marker_colour)
        # Create a custom formatted HTML table for each site marker and its popup
        site_html = populate_each_html_table_row_popup(site_details)
        site_popup = folium.Popup(
            folium.Html(site_html, script=True), max_width=1200, max_height=600
        )

        # For scrollbars, try this: # Reference: https://stackoverflow.com/a/73143192
        # Make sure your iFrame width matches the max width of the folium popup
        # iframe = branca.element.IFrame(html=site_html, width=710, height=300)
        # site_popup = folium.Popup(iframe, max_width=710, max_height=650)

        folium.Marker(
            location=(site_details["Latitude"], site_details["Longitude"]),
            icon=styled_icon,
            tooltip=site_details["Site name"],
            popup=site_popup,
        ).add_to(marker_layer)

    marker_layer.add_to(folium_map)

    return folium_map


def run_first_stage(cols: list[int]) -> pandas.DataFrame:
    """Run the first stage of the pipeline starting with reading dataset and ending with persisting internal state."""
    logger.info("Running first stage of the pipeline.")
    hld_df = read_dataset_to_df(
        dataset_path=QUALIFIED_DATASET_FILE, sheet_name="Sites", cols=cols
    )
    hld_df_filtered = filter_dataset(
        hld_df=hld_df,
        filter_column_name="New Update CE Property Jan 2023?",
        filter_criteria=["Yes", "Adjacent"],
        combination_operator=operator.or_,
    )
    hld_df_filtered_enriched = get_lat_long_postcode_from_easting_and_northing(
        hld_df=hld_df_filtered
    )
    hld_df_filtered_enriched_reordered = reorder_df_columns(
        hld_df=hld_df_filtered_enriched
    )
    save_intermediate_state(
        hld_df=hld_df_filtered_enriched_reordered,
        file_path=QUALIFIED_INTERMEDIATE_PICKLE_FILE,
    )

    logger.info("Finished first stage...")
    return hld_df_filtered_enriched_reordered


def run_second_stage(hld_df: pandas.DataFrame) -> folium.Map:
    """Run the second stage of the pipeline by plotting HCL site markers on an instance of OpenStreetMap."""
    logger.info("Running second stage of the pipeline.")

    folium_map = create_initial_folium_map(hld_df[["Latitude", "Longitude"]])
    hld_df_on_ce_property = hld_df[hld_df["New Update CE Property Jan 2023?"] == "Yes"]
    hld_df_adjacent_ce_property = hld_df[
        hld_df["New Update CE Property Jan 2023?"] == "Adjacent"
    ]

    # Plot site markers for sites that are on CE property
    folium_map = plot_site_markers_on_map(
        hld_df=hld_df_on_ce_property,
        marker_colour="red",
        marker_layer_name="On CE Property",
        folium_map=folium_map,
    )

    # Plot site markers for sites that are adjacent to CE property
    folium_map = plot_site_markers_on_map(
        hld_df=hld_df_adjacent_ce_property,
        marker_colour="blue",
        marker_layer_name="Adjacent to CE Property",
        folium_map=folium_map,
    )

    # Add the layer control
    folium.LayerControl().add_to(folium_map)

    logger.info("Finished second stage...")
    return folium_map


def run_programme(load_existing: bool = True) -> tuple[list[int], pandas.DataFrame]:
    """Main entry point for the whole programme to plot HCL site markers on the map from the dataset."""
    column_headers, column_letters, column_indices = get_column_letters_from_names(
        dataset_path=QUALIFIED_DATASET_FILE, sheet_index=0
    )
    (
        useful_column_letters,
        useful_column_nums,
    ) = convert_useful_col_names_to_col_letters_and_indices(
        column_headers, column_letters, column_indices
    )

    if load_existing:
        if not QUALIFIED_INTERMEDIATE_PICKLE_FILE.exists():
            hld_df_filtered_enriched_reordered = run_first_stage(useful_column_nums)

        else:
            hld_df_filtered_enriched_reordered = read_intermediate_state(
                file_path=QUALIFIED_INTERMEDIATE_PICKLE_FILE
            )

    else:  # Run the first stage regardless of pre-existing intermediate state
        hld_df_filtered_enriched_reordered = run_first_stage(useful_column_nums)

    folium_map = run_second_stage(hld_df_filtered_enriched_reordered)

    logger.info(f"Saving the final map to: {QUALIFIED_FOLIUM_MAP_FILE}")
    folium_map.save(f"{FOLIUM_MAP_FILE_NAME}")

    return useful_column_nums, hld_df_filtered_enriched_reordered


# Press the green button in the gutter to run the script.
if __name__ == "__main__":
    """Main entry point for reading HCL dataset and plotting appropriate site markers on an interactive map."""

    start_time = time.time()
    # useful_cols_nums, hld_df_dataset = run_programme(load_existing=False)  # Usually takes 00:02:33 minutes (24x slower)  # noqa
    useful_cols_nums, hld_df_dataset = run_programme(
        load_existing=True
    )  # Usually takes 00:00:02 seconds

    end_time = time.time()
    elapsed_time = end_time - start_time
    logger.info(
        f"All done! Programme execution took: {time.strftime('%H:%M:%S', time.gmtime(elapsed_time))}"
    )
